import openpyxl

def readdatafromexcel(path):
    "This function is just for reading the no of columns, rows in the excel file and also the Object of the file to read the data"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    rows = sheet_obj.max_row
    column = sheet_obj.max_column
    return sheet_obj,rows,column