import openpyxl
from Library import ConfigReader

def capture_tc_pass():
    tc_status = "NPassed"
    obj_wbtestresult = openpyxl.Workbook()
    sheettestresult = obj_wbtestresult.active
    c1 = sheettestresult.cell(row=1, column=1)
    c1.value = "Registration Test"
    c2 = sheettestresult.cell(row=1, column=2)
    c2.value = tc_status
    obj_wbtestresult.save(ConfigReader.readConfigdata("Test_Result_File", "registration"))

def capture_tc_failed():
    tc_status = "NFailed"
    obj_wbtestresult = openpyxl.Workbook()
    sheettestresult = obj_wbtestresult.active
    c1 = sheettestresult.cell(row=1, column=1)
    c1.value = "Registration Test"
    c2 = sheettestresult.cell(row=1, column=2)
    c2.value = tc_status
    obj_wbtestresult.save(ConfigReader.readConfigdata("Test_Result_File","registration"))