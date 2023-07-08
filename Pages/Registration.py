import logging
from Base import InitiateDriver
from Library import ConfigReader
from Base import Screenshot
from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.by import By



def registration(driver, sheet_obj,rows,column,name):

    for i in range(2, rows+1):
        for j in range(1, column+1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            data = cell_obj.value
            if j==1:
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_firstname")).send_keys(data)
            if j==2:
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_lastname")).send_keys(data)
            if j==3:
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_email")).send_keys(data)
            if j==4:
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_phoneno")).send_keys(data)
            if j==5:
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_pass")).send_keys(data)
            if j==6:
                driver.find_element(By.XPATH,ConfigReader.readConfigdata("Registration_Locators","signup_confirmpass")).send_keys(data)
                time.sleep(5)
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_social")).click()
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_accept")).click()
                driver.find_element(By.XPATH, ConfigReader.readConfigdata("Registration_Locators","signup_button")).click()
                time.sleep(5)
                Screenshot.take_page_screenshot(driver,name)
                driver.refresh()


def capture_tc_pass():
    tc_status = "Passed"
    obj_wbtestresult = openpyxl.Workbook()
    sheettestresult = obj_wbtestresult.active
    c1 = sheettestresult.cell(row=1, column=1)
    c1.value = "Registration Test"
    c2 = sheettestresult.cell(row=1, column=2)
    c2.value = tc_status
    obj_wbtestresult.save(ConfigReader.readConfigdata("Test_Result_File", "registration"))


def capture_tc_failed():
    tc_status = "Failed"
    obj_wbtestresult = openpyxl.Workbook()
    sheettestresult = obj_wbtestresult.active
    c1 = sheettestresult.cell(row=1, column=1)
    c1.value = "Registration Test"
    c2 = sheettestresult.cell(row=1, column=2)
    c2.value = tc_status
    obj_wbtestresult.save(ConfigReader.readConfigdata("Test_Result_File","registration"))